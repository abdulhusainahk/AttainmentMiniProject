from django.contrib.admin.helpers import AdminForm
from django.db.models import Avg
from django.shortcuts import render, redirect
from temp.tem.forms import EmployeeForm, DmForm, LoginForm, COForm, MappingForm, PosForm, \
    AdminLoginForm, TsForm, EsForm, WeightForm, TeacherForm, TeachesForm
from temp.tem.models import Employee, PO, CO, Upload_Int, Upload_Ext, ExamScheme, Assesment, Admin, Weights, CoPoMapp, \
    Report, DeliveryMethods, TeachingScheme
from django.http import HttpResponse
from django.views.generic import View
from .utils import render_to_pdf
from django.views.decorators.cache import cache_control
import openpyxl


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def blog(request):
    return render(request, "blog.html")


def courses1(request):
    return render(request, "courses1.html")


def dbms(request):
    return render(request, "dbms.html")


def home(request):
    return render(request, "home.html")


def programs(request):
    return render(request, "programs.html")


def os(request):
    return render(request, "os.html")


def hci(request):
    return render(request, "hci.html")


def toc(request):
    return render(request, "toc.html")


def sepm(request):
    return render(request, "sepm.html")


global flag
flag = 0


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def showco(request):
    s1 = request.session['course']  # login code and pattern e.g 1441_2015
    global v1, cnt, s2
    s2 = s1
    print(s2)
    if request.method == "POST":
        form = COForm(request.POST)
        if form.is_valid:
            try:
                temp = form.save(commit=False)  # creating temporary object
                s2 = s2 + '.' + str(cnt)  # 1441_2015.1 ,.2, .3 .....
                temp.CO_id = s2  # assign that value
                temp.save()
                p = CoPoMapp(CO_id_id=s2, VPO1=0, VPO2=0, VPO3=0, VPO4=0, VPO5=0, VPO6=0, VPO7=0, VPO8=0, VPO9=0,
                             VPO10=0, VPO11=0, VPO12=0)
                p.save()  # save corrosponding PO directly
                try:
                    weightsobj = Weights.objects.get(Code_patt_id=s1)
                    calcresult(temp, s1, 1)
                except:
                    pass
                return redirect('/course')
            except:
                print('in except')
                pass
        else:
            print('not validated')
    else:
        cnt = 1  # initializing first counter
        co = CO.objects.filter(CO_id__icontains=s1)
        for i in co:
            cnt = cnt + 1  # counter from which CO will be entered
        form = COForm()
        v1 = s2 + '.' + str(cnt)
    return render(request, "newco.html", {'form': form, 'v1': v1})


def course(request):
    s1 = request.session['course']
    po = []
    try:
        co = CO.objects.filter(CO_id__icontains=s1)
        for i in co:
            p = CoPoMapp.objects.get(CO_id_id=i.CO_id)
            po.append(p)
    except:
        co = []
        po = []
    return render(request, "courses.html", {'co': co, 'po': po})


def edit(request, id):
    obj = CO.objects.get(CO_id=id)
    form = COForm(
        initial={'Outcome': obj.Outcome, 'assignments': obj.assignments, 'units': obj.units, 'insem': obj.insem,
                 'endsem': obj.endsem, 'practicle': obj.practicle, 'TW': obj.TW})
    return render(request, "edit.html", {'obj': obj, 'form': form})


def update(request, id):
    obj = CO.objects.get(CO_id=id)
    form = COForm(request.POST)
    if form.is_valid():
        temp = form.save(commit=False)
        temp.CO_id = id
        # map this co to po's here itself
        temp.save()
        assessment_obj = Assesment.objects.get(CO_id_id=id)
        assessment_obj.delete()
        calcresult(temp, request.session['course'], 1)
        return redirect('/course')
    form = COForm(
        initial={'Outcome': obj.Outcome, 'assignments': obj.assignments, 'units': obj.units, 'insem': obj.insem,
                 'endsem': obj.endsem, 'practicle': obj.practicle, 'TW': obj.TW})
    return render(request, "edit.html", {'obj': obj, 'form': form})


def delete(request, id):
    global t2
    s1 = request.session['course']
    temp = int(id[-1])  # Which id will be deleted
    coobj = CO.objects.get(CO_id=id)
    coobj.delete()
    changeid(0, temp, s1)  # for co
    return redirect('/course')


def changeid(flag, temp, s1):
    co = CO.objects.filter(CO_id__icontains=s1)
    for i in co:
        t2 = int(i.CO_id[-1])
        if t2 > temp:
            t2 = t2 - 1
            newid = s1 + '.' + str(t2)  # new co i.e if 2 deleted then 3 becomes 2
            if flag == 1:
                CoPoMapp.objects.filter(CO_id_id=i.CO_id).update(CO_id_id=newid)
            else:
                CO.objects.filter(CO_id=i.CO_id).update(CO_id=newid)
    return


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def mapping(request):
    s1 = request.session['course']
    q = []
    q2 = []
    flag = 0
    co = CO.objects.filter(CO_id__icontains=s1)
    for j in co:
        try:
            obj = CoPoMapp.objects.get(CO_id=j)
            if obj.VPO1 >= 0:
                q2.append(obj)
                flag = 1
            else:
                return render(request, "blog.html", {'q': q2})
        except:
            for i in co:
                p = CoPoMapp(CO_id_id=i.CO_id, VPO1=0, VPO2=0, VPO3=0, VPO4=0, VPO5=0, VPO6=0, VPO7=0, VPO8=0, VPO9=0,
                             VPO10=0, VPO11=0, VPO12=0)
                q.append(p)
                p.save()
            return render(request, "blog.html", {'q': q})
    return render(request, "blog.html", {'q': q2})


def editmap(request, id):
    obj = CoPoMapp.objects.get(CO_id_id=id)
    coobj = CO.objects.get(CO_id=id)
    form = MappingForm(
        initial={'VPO1': obj.VPO1, 'VPO2': obj.VPO2, 'VPO3': obj.VPO3, 'VPO4': obj.VPO4, 'VPO5': obj.VPO5,
                 'VPO6': obj.VPO6, 'VPO7': obj.VPO7, 'VPO8': obj.VPO8, 'VPO9': obj.VPO9, 'VPO10': obj.VPO10,
                 'VPO11': obj.VPO11, 'VPO12': obj.VPO12})
    return render(request, "editmap.html", {'obj': obj, 'form': form, 'coobj': coobj})


def updatemap(request, id):
    coobj = CO.objects.get(CO_id=id)
    obj = CoPoMapp.objects.get(CO_id_id=id)
    form = MappingForm(request.POST)
    if form.is_valid():
        temp = form.save(commit=False)
        temp.CO_id = coobj
        temp.save()
        return redirect('/course')
    form = MappingForm(
        initial={'VPO1': obj.VPO1, 'VPO2': obj.VPO2, 'VPO3': obj.VPO3, 'VPO4': obj.VPO4, 'VPO5': obj.VPO5,
                 'VPO6': obj.VPO6, 'VPO7': obj.VPO7, 'VPO8': obj.VPO8, 'VPO9': obj.VPO9, 'VPO10': obj.VPO10,
                 'VPO11': obj.VPO11, 'VPO12': obj.VPO12})
    return render(request, "edit.html", {'obj': obj, 'form': form, 'coobj': coobj})


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def contact(request):
    v = request.session['course']
    print(v)
    try:
        obj = Weights.objects.get(Code_patt_id=v)
    except:
        obj = []
    form = WeightForm()
    print(obj)
    return render(request, "contact.html", {'v': v, 'form': form, 'obj': obj})


def pricing(request):
    return redirect('/show')


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def index1(request):
    return render(request, "index1.html")


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def emp(request):
    if request.method == "POST":
        if request.POST.get('password') == request.POST.get('re-password'):
            form = EmployeeForm(request.POST)
            request.session['course'] = request.POST.get('course')
            request.session['password'] = request.POST.get('password')
            if form.is_valid():
                try:
                    form.save()
                    return redirect('/access')
                except:
                    pass
            else:
                print(1)
                return redirect('/')
        else:
            print(2)
            return redirect('/')
    else:
        form = EmployeeForm()
        return render(request, "home.html", {'form': form})


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def loginpage(request):
    data2 = Employee.objects.all()
    if request.method == "POST":
        request.session['course'] = request.POST.get('course')
        request.session['password'] = request.POST.get('password')
        if request.POST.get('password') == request.POST.get('re-password'):
            form1 = LoginForm(request.POST)
            if form1.is_valid():
                print(3)
                ext = form1.cleaned_data
                for i in data2:
                    codev = ext.get("course")
                    paswd = ext.get("password")
                    if i.course == codev and i.password == paswd:
                        return redirect('/access')

                return redirect('/')
        else:
            return redirect('/')
    else:
        form1 = EmployeeForm()
        return render(request, "index.html", {'form1': form1})


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def access_session(request):
    data1 = {
        'x': request.session['course'],
        'y': request.session['password']
    }
    form = DmForm()
    try:
        obj = DeliveryMethods.objects.get(Code_patt=data1['x'])
        print(obj.ict)
        flag = 1
    except:
        obj = []
        flag = 0
        return render(request, "index1.html", {'data1': data1, 'form': form, 'obj': obj, 'flag': flag})
    return render(request, "index1.html", {'data1': data1, 'form': form, 'obj': obj, 'flag': flag})


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def dm(request):  # -------------logic of dmform with foreign keys
    global v
    if request.method == "POST":
        form = DmForm(request.POST)
        if form.is_valid():
            try:
                s1 = request.session['course']  # 1441_2015 or any code like that coming from login
                obj2 = Employee.objects.get(course=s1)  # object where Code_patt == logined code
                obj = form.save(commit=False)
                try:
                    temp = DeliveryMethods.objects.get(Code_patt_id=s1)  # check for update
                    DeliveryMethods.objets.filter(Code_patt_id=s1).update(chalktalk=obj.chalktalk, ict=obj.ict,
                                                                          gd=obj.gd, ifv=obj.ifv, et=obj.et,
                                                                          sur=obj.sur, mp=obj.mp, lab=obj.lab)
                except:
                    obj.Code_patt = obj2  # assign that value
                    obj.save()  # save it

                print('saved')
                return redirect('/dm')
            except:
                print('unsaved')
                return redirect('/dm')
                pass
        else:
            print('not validated')
    else:
        form = DmForm()
        s1 = request.session['course']
        obj = Employee.objects.get(course=s1)
        v = obj  # passing value of the code you have logged in we can also use session handling here
    return redirect('/access')


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def show(request):
    pos = {
        'pos1': PO.objects.get(id="1"),
        'pos2': PO.objects.get(id="2"),
        'pos3': PO.objects.get(id="3"),
        'pos4': PO.objects.get(id="4"),
        'pos5': PO.objects.get(id="5"),
        'pos6': PO.objects.get(id="6"),
        'pos7': PO.objects.get(id="7"),
        'pos8': PO.objects.get(id="8"),
        'pos9': PO.objects.get(id="9"),
        'pos10': PO.objects.get(id="10"),
        'pos11': PO.objects.get(id="11"),
        'pos12': PO.objects.get(id="12")
    }
    s1 = request.session['course']
    v = s1
    return render(request, "pricing.html", {'pos': pos, 'v': v})


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def delete_session(request):
    try:
        del request.session['course']
        del request.session['password']
    except KeyError:
        pass
    return redirect('/')


class GeneratePdf(View):
    def get(self, request, *args, **kwargs):
        print(2)
        s1 = request.session['course']
        coursedetails = Employee.objects.get(course=s1)
        DM = DeliveryMethods.objects.filter(Code_patt=s1)
        cos = CO.objects.filter(CO_id__icontains=s1)
        pos = PO.objects.all()
        # map = CoPoMapp.objects.filter(CO_id_id__icontains=s1)
        data = {
            'coursedetails': coursedetails,
            'DM': DM,
            'CO': cos,
            'po': pos,
        }
        print(3)
        pdf = render_to_pdf('pdf/invoice.html', data)
        print(4)
        if pdf:
            response = HttpResponse(pdf, content_type='application/pdf')
            filename = "Invoice_%s.pdf" % ("12341231")
            content = "inline; filename='%s'" % (filename)
            download = request.GET.get("download")
            if download:
                content = "attachment; filename='%s'" % (filename)
            response['Content-Disposition'] = content
            return response
        return HttpResponse("Not found")


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def ind(request):
    if "POST" == request.method:
        excel_file = request.FILES["excel_file"]
        wb = openpyxl.load_workbook(excel_file)
        sheets = wb.sheetnames
        worksheet = wb["Sheet1"]
        active_sheet = wb.active
        excel_data = list()
        arr = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
        for row in worksheet.iter_rows():
            row_data = list()
            cnt = 0
            for cell in row:
                row_data.append(str(cell.value))
                arr[cnt] = cell.value
                cnt = cnt + 1
            x = request.session['course']
            obj1 = Employee.objects.get(course=x)
            p = Upload_Int(Code_patt=obj1, A1=arr[0], A2=arr[1], A3=arr[2], A4=arr[3], A5=arr[4], ut1=arr[5],
                           ut2=arr[6], ut3=arr[7], ut4=arr[8], ut5=arr[9], tw=arr[10])
            p.save()
            excel_data.append(row_data)

        # form.save()
        return redirect('/contact')
    else:
        return render(request, 'contact.html', {})


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def indi(request):
    if "POST" == request.method:
        excel_file = request.FILES["excel_file"]
        wb = openpyxl.load_workbook(excel_file)
        sheets = wb.sheetnames
        worksheet = wb["Sheet1"]
        active_sheet = wb.active
        excel_data = list()
        arr = [0, 0, 0, 0, 0]
        for row in worksheet.iter_rows():
            row_data = list()
            cnt = 0
            for cell in row:
                row_data.append(str(cell.value))
                arr[cnt] = cell.value
                cnt = cnt + 1

            x1 = request.session['course']
            obj2 = Employee.objects.get(course=x1)
            p = Upload_Ext(Code_patt=obj2, insem=arr[0], practicals=arr[1], TW=arr[2], endsem=arr[3])
            p.save()
            excel_data.append(row_data)

        # form.save()
        return redirect('/contact')
    else:
        return render(request, 'contact.html', {})


def calcresult(obj, code, flag):
    assign = []
    units = []
    avgassn = 0
    avgunits = 0
    avgtw = 0
    avgpract = 0
    avginsem = 0
    avgendsem = 0
    intcount = extcount = 0
    for i in obj.assignments:
        if i != ',':
            assign.append(int(i))
    for i in obj.units:
        if i != ',':
            units.append(int(i))
    sum = count = 0
    for i in assign:
        count = count + 1
        assnno = "A" + str(i)
        x = Upload_Int.objects.filter(Code_patt=code).aggregate(Avg(assnno))
        intcount = intcount + 1
        assnno = assnno + '__avg'
        sum = (sum + (x[assnno]))
    avgassn = sum / count
    sum = count = 0
    for i in units:
        count = count + 1
        unitno = "ut" + str(i)
        x = Upload_Int.objects.filter(Code_patt=code).aggregate(Avg(unitno))
        intcount = intcount + 1
        unitno = unitno + '__avg'
        sum = (sum + (x[unitno]))

    avgunits = sum / count
    es = ExamScheme.objects.get(Code_patt=code)
    if obj.TW:
        x = Upload_Int.objects.filter(Code_patt=code).aggregate(Avg('tw'))
        intcount = intcount + 1
        avgtw = 100 * x['tw__avg'] / es.termwork
    if obj.practicle:
        x = Upload_Ext.objects.filter(Code_patt=code).aggregate(Avg('practicals'))
        extcount = extcount + 1
        avgpract = 100 * x['practicals__avg'] / es.Practicle
    if obj.insem:
        x = Upload_Ext.objects.filter(Code_patt=code).aggregate(Avg('insem'))
        extcount = extcount + 1
        avginsem = 100 * x['insem__avg'] / es.onlineinsem
    if obj.endsem:
        x = Upload_Ext.objects.filter(Code_patt=code).aggregate(Avg('endsem'))
        extcount = extcount + 1
        avgendsem = 100 * x['endsem__avg'] / es.endsem

    weight = Weights.objects.get(Code_patt=code)
    attn = (weight.int_weight * (avgunits + avgassn + avgtw) / intcount) + (
            weight.ext_weight * (avgpract + avgendsem + avginsem) / extcount)
    if attn >= 80:
        lvl = 3
    elif attn >= 70:
        lvl = 2
    elif attn >= 60:
        lvl = 1
    else:
        lvl = 0
    if flag == 1:
        # print('saving')
        p = Assesment(CO_id=obj, assignment=avgassn, insem=avginsem, endsem=avgendsem, practicle=avgpract,
                      termwork=avgtw, unittests=avgunits, assessment=attn, level=lvl)
        p.save()
    else:
        # print('updating')
        Assesment.objects.filter(CO_id=obj).update(CO_id=obj, assignment=avgassn, insem=avginsem, endsem=avgendsem,
                                                   practicle=avgpract,
                                                   termwork=avgtw, unittests=avgunits, assessment=attn, level=lvl)
        # print('updated')
        flag = 0
    return


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def result(request):
    code = request.session['course']  # 1441(2015)
    cos = CO.objects.filter(CO_id__icontains=code)
    for i in cos:
        try:
            done = Assesment.objects.get(CO_id=i)  # already calculated results
        except:
            calcresult(i, code, 1)
    return redirect('/display2')


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def display2(request):
    newform = []
    s1 = request.session['course']
    form = Assesment.objects.all()
    for i in form:
        if i.CO_id_id[:-2] == s1:
            newform.append(i)
    return render(request, "report.html", {'newform': newform, 's1': s1})


def courseinfo(request):
    if request.method == "POST":
        form2 = EmployeeForm(request.POST)
        if form2.is_valid():
            try:
                form2.save()
                return redirect('/adminhomepage')
            except:
                pass
    else:
        form2 = EmployeeForm()
    return render(request, "entry.html", {'form2': form2})


def adminhomepage(request):
    return render(request, "adminhome.html")


def adminentry(request):
    if request.method == "POST":
        form = AdminForm(request.POST)
        if form.is_valid():
            try:
                form.save()
                return redirect('/adminlogin')
            except:
                pass
    else:
        form = AdminForm()
    return render(request, "adminentry.html", {'form': form})


def adminlogin(request):
    if request.method == "POST":
        form = AdminLoginForm(request.POST)
        data2 = Admin.objects.all()
        if form.is_valid():
            ext = form.cleaned_data
            for i in data2:
                uname = ext.get("username")
                paswd = ext.get("password")
                if i.username == uname and i.password == paswd:
                    return redirect('/adminhomepage')
                else:
                    return redirect('/adminentry')

            return redirect('/')
        else:
            form = AdminLoginForm()
            return render(request, "adinlogin.html", {'form': form})


def poentry(request):
    if request.method == "POST":
        form = PosForm(request.POST)
        if form.is_valid:
            try:
                form.save()
                form = PosForm()
                return render(request, "PO.html", {'form': form})
            except:
                print('unsaved')
                pass
        else:
            print('not validated')
    else:
        form = PosForm()
    return render(request, "PO.html", {'form': form})


def coursescheme(request):
    global s1
    s1 = Employee.ename
    if request.method == "POST":
        form = TsForm(request.POST)
        form2 = EsForm(request.POST)
        if form.is_valid() and form2.is_valid():
            try:
                f1 = form.save(commit=False)
                f2 = form2.save(commit=False)
                s1 = Employee.ename  # 1441_2015 or any code like that coming from login
                obj2 = Employee.objects.get(Code_patt=s1)
                f1.Code_patt = obj2
                f2.Code_patt = obj2
                f1.save()
                f2.save()
                print('hhhh')
                return redirect('/homepage2')
            except:
                print('unsaved')
                pass
        else:
            print('not validated')
    else:
        form = TsForm()
        form2 = EsForm()
    print(s1)
    return render(request, "schemes.html", {'form': form, 'form2': form2, 's1': s1})


def weights(request):
    if request.method == "POST":
        form = WeightForm(request.POST)
        if form.is_valid:
            try:
                s1 = request.session['course']  # 1441_2015 or any code like that coming from login
                obj2 = Employee.objects.get(course=s1)
                obj = form.save(commit=False)
                x = obj.ext_weight
                y = 1 - x
                y = y * 10
                y = round(y) / 10
                try:
                    temp = Weights.objects.get(Code_patt_id=s1)
                    Weights.objects.filter(Code_patt_id=s1).update(int_weight=y, ext_weight=x, target=obj.target)
                    flag = 1
                except:
                    obj.int_weight = y
                    obj.Code_patt = obj2
                    obj.save()
                return redirect('/contact')
            except:
                print('unsaved')
                pass
        else:
            print('not validated')

    return redirect('/contact')


def teacherentry(request):
    if request.method == "POST":
        form2 = TeacherForm(request.POST)
        if form2.is_valid():
            try:
                form2.save()
                return redirect('/adminhomepage')
            except:
                print('unsaved')
                pass
    else:
        print('not validated')
        form2 = TeacherForm()
    return render(request, "teacherentry.html", {'form2': form2})


def teacher_alloc(request):
    if request.method == "POST":
        form2 = TeachesForm(request.POST)
        if form2.is_valid():
            try:
                form2.save()
                return redirect('/adminhomepage')
            except:
                pass
    else:
        form2 = TeachesForm()
    return render(request, "teacher_alloc.html", {'form2': form2})


def report_gen(request):
    # Co po mapping , assessment
    code = request.session['course']
    results = Assesment.objects.all()
    mappings = CoPoMapp.objects.all()
    for i in results:
        print(i.CO_id_id[:-2], code)
        if i.CO_id_id[:-2] == code:  # got assessment of a CO of logged in course
            print('yess')
            for j in mappings:
                if j.CO_id_id == i.CO_id_id:
                    r1 = (i.assignment * j.VPO1) / 3
                    r2 = (i.assignment * j.VPO2) / 3
                    r3 = (i.assignment * j.VPO3) / 3
                    r4 = (i.assignment * j.VPO4) / 3
                    r5 = (i.assignment * j.VPO5) / 3
                    r6 = (i.assignment * j.VPO6) / 3
                    r7 = (i.assignment * j.VPO7) / 3
                    r8 = (i.assignment * j.VPO8) / 3
                    r9 = (i.assignment * j.VPO9) / 3
                    r10 = (i.assignment * j.VPO10) / 3
                    r11 = (i.assignment * j.VPO11) / 3
                    r12 = (i.assignment * j.VPO12) / 3

                    cd = Employee.objects.get(course=code)
                    name = cd.CourseName
                    check = Weights.objects.get(Code_patt_id=code)
                    if i.assignment >= check.target:
                        yn = 'YES'
                    else:
                        yn = 'NO'
                    b = Report(Course_name=name, assessment=i.assessment, CO_id=j.CO_id, Y_N=yn, VPO1=r1, VPO2=r2,
                               VPO3=r3, VPO4=r4, VPO5=r5,
                               VPO6=r6, VPO7=r7, VPO8=r8, VPO9=r9, VPO10=r10, VPO11=r11, VPO12=r12)
                    b.save()

    return redirect('/showreport')


def showreport(request):
    s1 = request.session['course']
    obj = []
    form = Report.objects.all()
    for i in form:
        if i.CO_id_id[:-2] == s1:
            obj.append(i)
    print(obj)
    return render(request, "report.html", {'obj': obj, 's1': s1})


def gencis(request):
    global target
    obj = []
    res = []
    k = 0
    s1 = request.session['course']
    code = s1[:4]
    pattern = s1[5:-1]
    course = Employee.objects.get(course=s1)

    print(s1)
    dm = DeliveryMethods.objects.get(Code_patt=s1)
    wt = Weights.objects.filter(Code_patt=s1)
    for z in wt:
        target = z.target
    a = b = c = d = e = f = g = h = 0
    if dm.chalktalk:
        a = 1
    if dm.ict:
        b = 1
    if dm.gd:
        c = 1
    if dm.ifv:
        d = 1
    if dm.et:
        e = 1
    if dm.sur:
        f = 1
    if dm.mp:
        g = 1
    if dm.lab:
        h = 1
    data = {
        'a': a,
        'b': b,
        'c': c,
        'd': d,
        'e': e,
        'f': f,
        'g': g,
        'h': h,
    }
    cnt = 0
    co = CO.objects.filter(CO_id__icontains=s1)
    for x in co:
        cnt = cnt + 1

    pos = PO.objects.all()
    map = CoPoMapp.objects.all()
    ass = Assesment.objects.all()
    j = 1
    for i in map:
        if j > cnt:
            break
        cmp = s1 + '.' + str(j)
        if i.CO_id_id == cmp:
            obj.append(i)
            j = j + 1

    j = 1
    for i in ass:
        if j > cnt:
            break
        cmp = s1 + '.' + str(j)
        if i.CO_id_id == cmp:
            res.append(i)
            print(i.level)
            j = j + 1
    result1 = {
        'res': res,
        'target': target,
    }
    es = ExamScheme.objects.get(Code_patt=s1)
    ts = TeachingScheme.objects.get(Code_patt=s1)
    return render(request, "cis.html",
                  {'data': data, 'course': course, 'result1': result1, 'obj': obj, 'co': co, 'pos': pos, 'es': es,
                   'ts': ts, 'code': code, 'pattern': pattern})
